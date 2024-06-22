import sys
import threading
import subprocess
import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QTextEdit, QVBoxLayout, QWidget, QFileDialog, QInputDialog

# Lista de bibliotecas necesarias
LIBRERIAS_REQUERIDAS = ["PyQt5", "openpyxl"]

# Función para verificar si las bibliotecas necesarias están instaladas
def verificar_bibliotecas():
    for libreria in LIBRERIAS_REQUERIDAS:
        if libreria not in sys.modules:
            return False
    return True

# Función para instalar bibliotecas faltantes utilizando pip
def instalar_bibliotecas_faltantes():
    for libreria in LIBRERIAS_REQUERIDAS:
        subprocess.run([sys.executable, "-m", "pip", "install", libreria])

# Función de desencriptación de mensajes VITROS (ejemplo)
def desencriptar_mensaje(mensaje):
    # Aquí va la lógica de desencriptación (ejemplo: convertir a mayúsculas)
    return mensaje.upper()

# Función de ordenación de mensajes VITROS (ejemplo)
def ordenar_mensajes(mensajes):
    # Aquí va la lógica de ordenación (ejemplo: ordenar alfabéticamente)
    return sorted(mensajes)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Captura y decodificación de mensajes VITROS ECiQ")
        self.setGeometry(100, 100, 600, 400)

        # Verificar si las bibliotecas necesarias están instaladas
        if not verificar_bibliotecas():
            print("Algunas bibliotecas necesarias no están instaladas. Intentando instalarlas...")
            instalar_bibliotecas_faltantes()
            # Verificar nuevamente después de la instalación
            if not verificar_bibliotecas():
                print("No se pudieron instalar todas las bibliotecas necesarias.")
                sys.exit(1)

        # Detalles de la conexión
        self.lbl_estado = QLabel("Estado: Desconectado")
        self.lbl_inicio = QLabel("Inicio: -")
        self.lbl_dispositivo = QLabel("Dispositivo: -")
        self.lbl_puerto_serie = QLabel("Puerto Serie: -")

        # Crear un área de texto para mostrar los mensajes
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)

        # Botones para conexión y desconexión
        self.btn_conectar = QPushButton("Conectar")
        self.btn_conectar.clicked.connect(self.conectar_dispositivo)

        self.btn_desconectar = QPushButton("Desconectar")
        self.btn_desconectar.clicked.connect(self.desconectar_dispositivo)
        self.btn_desconectar.setEnabled(False)

        self.btn_generar_excel = QPushButton("Generar Excel")
        self.btn_generar_excel.clicked.connect(self.generar_excel_dialogo)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.lbl_estado)
        layout.addWidget(self.lbl_inicio)
        layout.addWidget(self.lbl_dispositivo)
        layout.addWidget(self.lbl_puerto_serie)
        layout.addWidget(self.text_edit)
        layout.addWidget(self.btn_conectar)
        layout.addWidget(self.btn_desconectar)
        layout.addWidget(self.btn_generar_excel)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        # Variables para la comunicación Kermit
        self.proceso_kermit = None
        self.hilo_lectura = None
        self.conectado = False
        self.datos_kermit = []

    # Función para conectar el dispositivo
    def conectar_dispositivo(self):
        try:
            # Actualizar información del dispositivo y puerto serie
            dispositivo = "Dispositivo: Comunicaciones3"
            puerto_serie = "Puerto Serie: Com1 9600, 8, None, 1"
            self.lbl_dispositivo.setText(dispositivo)
            self.lbl_puerto_serie.setText(puerto_serie)

            # Comando para iniciar Kermit y establecer una conexión
            comando_kermit = "kermit -l COM1 -b 9600"

            # Iniciar sesión Kermit
            self.proceso_kermit = subprocess.Popen(comando_kermit, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            self.conectado = True

            # Actualizar la interfaz
            self.lbl_estado.setText("Estado: Conectado")
            self.btn_conectar.setEnabled(False)
            self.btn_desconectar.setEnabled(True)

            # Iniciar el hilo para la lectura de mensajes
            self.hilo_lectura = threading.Thread(target=self.leer_mensajes)
            self.hilo_lectura.start()

        except Exception as e:
            self.text_edit.append("Error al conectar el dispositivo: " + str(e))

    # Función para desconectar el dispositivo
    def desconectar_dispositivo(self):
        if self.conectado:
            # Detener el hilo de lectura
            self.conectado = False
            self.hilo_lectura.join()

            # Cerrar sesión Kermit
            self.proceso_kermit.terminate()
            self.proceso_kermit = None

            # Actualizar la interfaz
            self.lbl_estado.setText("Estado: Desconectado")
            self.btn_conectar.setEnabled(True)
            self.btn_desconectar.setEnabled(False)

    # Función para generar el archivo Excel con diálogo de selección de examen y resultado
    def generar_excel_dialogo(self):
        examen, ok = QInputDialog.getText(self, "Examen", "Nombre del examen:")
        if ok:
            resultado, ok = QInputDialog.getText(self, "Resultado", "Resultado del examen:")
            if ok:
                self.generar_excel(examen, resultado)

    # Función para generar el archivo Excel con datos capturados por la sesión Kermit, el examen y su resultado
    def generar_excel(self, examen, resultado):
        try:
            # Desencriptar y ordenar los mensajes
            mensajes_desencriptados = [desencriptar_mensaje(m) for m in self.datos_kermit]
            mensajes_ordenados = ordenar_mensajes(mensajes_desencriptados)

            # Generar archivo Excel
            wb = openpyxl.Workbook()
            ws = wb.active

            # Escribir encabezados
            encabezados = ["ID", "Mensaje", "Examen", "Resultado"]
            for i, encabezado in enumerate(encabezados, start=1):
                ws.cell(row=1, column=i, value=encabezado)

            # Escribir datos de los mensajes desencriptados y ordenados
            for index, mensaje in enumerate(mensajes_ordenados, start=2):
                id_mensaje =  f"{index:02d}.{str(index).zfill(4)}"
                ws.cell(row=index, column=1, value=id_mensaje)
                ws.cell(row=index, column=2, value=mensaje)
                ws.cell(row=index, column=3, value=examen)
                ws.cell(row=index, column=4, value=resultado)

            # Mostrar diálogo de selección de archivo para elegir la ubicación y nombre del archivo Excel
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(self, "Guardar archivo Excel", "", "Archivos de Excel (*.xlsx)", options=options)

            # Guardar archivo Excel en la ubicación seleccionada por el usuario
            if file_path:
                wb.save(file_path)
                self.text_edit.append(f"Se ha guardado el archivo Excel en: {file_path}")

        except Exception as e:
            self.text_edit.append("Error al generar el archivo Excel: " + str(e))

    # Función para leer los mensajes en tiempo real
    def leer_mensajes(self):
        while self.conectado:
            mensaje = self.proceso_kermit.stdout.readline().decode('utf-8').strip()
            if mensaje:
                self.text_edit.append(mensaje)
                self.datos_kermit.append(mensaje)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

