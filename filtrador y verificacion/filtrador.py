import os
import subprocess
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QTreeWidget, QTreeWidgetItem, QMenu, QAction

# Función para instalar o actualizar las librerías necesarias
def instalar_actualizar_librerias():
    try:
        # Comando para instalar o actualizar las librerías
        subprocess.run([sys.executable, "-m", "pip", "install", "--upgrade", "pyqt5", "openpyxl"])
    except Exception as e:
        print(f"No se pudieron instalar/actualizar las librerías: {e}")

class AplicacionExamenMedico(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Filtrar Exámenes Médicos")
        self.setGeometry(100, 100, 600, 400)

        self.crear_interfaz()

    def crear_interfaz(self):
        # Widget principal
        widget_principal = QWidget()
        layout_principal = QVBoxLayout()
        widget_principal.setLayout(layout_principal)
        self.setCentralWidget(widget_principal)

        # Elementos de entrada
        self.texto_entry = QLineEdit()
        layout_principal.addWidget(QLabel("Texto a buscar en el nombre del examen:"))
        layout_principal.addWidget(self.texto_entry)

        # Botón para iniciar la búsqueda
        buscar_button = QPushButton("Buscar")
        buscar_button.clicked.connect(self.filtrar_examenes)
        layout_principal.addWidget(buscar_button)

        # Etiqueta para mostrar el estado de la búsqueda
        self.estado_label = QLabel("")
        layout_principal.addWidget(self.estado_label)

        # Resultados en TreeWidget
        self.resultados_treewidget = QTreeWidget()
        self.resultados_treewidget.setHeaderLabels(["Código Fonasa", "Nombre del Examen"])
        layout_principal.addWidget(self.resultados_treewidget)

    def cargar_base_datos(self, ruta_archivo):
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(ruta_archivo)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append((str(row[0]), str(row[1])))  # Convertir a cadena para evitar errores
            return data
        except Exception as e:
            print(f"No se pudo cargar la base de datos desde el archivo: {e}")
            return None

    def filtrar_examenes(self):
        texto_buscado = self.texto_entry.text().strip()
        if texto_buscado == "":
            self.estado_label.setText("Por favor ingrese un texto para buscar")
            return

        ruta_archivo = os.path.join(os.path.dirname(__file__), "maestrofonasa.xlsx")  # Ruta del archivo
        base_datos = self.cargar_base_datos(ruta_archivo)
        if base_datos is None:
            return

        resultados = [row for row in base_datos if texto_buscado.lower() in row[1].lower()]

        if resultados:
            self.mostrar_resultados(resultados)
            self.estado_label.setText("Búsqueda completada.")
        else:
            self.estado_label.setText(f"No se encontraron exámenes médicos que coincidan con '{texto_buscado}'")

    def mostrar_resultados(self, resultados):
        self.resultados_treewidget.clear()
        for row in resultados:
            item = QTreeWidgetItem(self.resultados_treewidget, [row[0], row[1]])

    def context_menu(self, pos):
        menu = QMenu()
        copiar_nombre_action = QAction("Copiar nombre", self)
        copiar_nombre_action.triggered.connect(self.copiar_nombre)
        menu.addAction(copiar_nombre_action)

        copiar_codigo_action = QAction("Copiar código", self)
        copiar_codigo_action.triggered.connect(self.copiar_codigo)
        menu.addAction(copiar_codigo_action)

        copiar_todo_action = QAction("Copiar todo", self)
        copiar_todo_action.triggered.connect(self.copiar_todo)
        menu.addAction(copiar_todo_action)

        menu.exec_(self.resultados_treewidget.viewport().mapToGlobal(pos))

    def copiar_nombre(self):
        item = self.resultados_treewidget.currentItem()
        if item:
            nombre = item.text(1)
            QApplication.clipboard().setText(nombre)

    def copiar_codigo(self):
        item = self.resultados_treewidget.currentItem()
        if item:
            codigo = item.text(0)
            QApplication.clipboard().setText(codigo)

    def copiar_todo(self):
        item = self.resultados_treewidget.currentItem()
        if item:
            nombre = item.text(1)
            codigo = item.text(0)
            texto = f"{codigo} - {nombre}"
            QApplication.clipboard().setText(texto)

if __name__ == "__main__":
    # Instalar o actualizar las librerías necesarias
    instalar_actualizar_librerias()

    # Crear la aplicación
    app = QApplication(sys.argv)
    ventana = AplicacionExamenMedico()
    ventana.show()
    sys.exit(app.exec_())
